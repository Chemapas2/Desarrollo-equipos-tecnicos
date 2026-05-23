
import io
import math
import base64
import tempfile
from datetime import datetime
from pathlib import Path
from html import escape

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage


st.set_page_config(
    page_title="TechTeam · Desarrollo técnico de equipos",
    page_icon="🧭",
    layout="wide",
)

APP_PASSWORD = "TechTeam2026+"

TRUNK_ORDER = ["Alimentación", "Sanidad", "Manejo", "Herramientas"]

MONOGASTRIC_SPECIES = ["Equipo mixto de monogástricos", "Avicultura", "Porcino", "Conejos", "Caballos"]
RUMINANT_SPECIES = ["Equipo mixto de rumiantes", "Vacuno leche", "Vacuno carne", "Ovino y caprino"]

TEAM_STATUS_RULES = {
    "Fortaleza consolidada": {"min_mean": 1.00, "min_refs": 2},
    "Aceptable con riesgo": {"min_mean": 0.95, "min_refs": 1},
    "Gap moderado": {"min_mean": 0.85, "min_refs": 0},
    "Gap crítico": {"min_mean": 0.00, "min_refs": 0},
}

STYLE = """
<style>
:root {
  --nutreco-blue:#143b8f;
  --techteam-red:#ef233c;
  --pink:#d81b90;
  --line:#dbe3ef;
  --muted:#6b7280;
  --soft:#f6f8fc;
}
.block-container {padding-top: 1rem; padding-bottom: 2rem;}
div[data-testid="stMetric"]{
  background:#ffffff;
  border:1px solid var(--line);
  border-radius:14px;
  padding:10px 14px;
}
div[data-testid="stTabs"] button {
  font-weight:700;
}
.stDownloadButton button, .stButton button {
  font-weight:700;
}
.small-note {
  color: var(--muted);
  font-size: 0.92rem;
}
.corp-hero {
  background: linear-gradient(135deg, rgba(20,59,143,.98), rgba(216,27,144,.92));
  border-radius: 22px;
  padding: 18px 22px;
  color: white;
}
.corp-chip {
  display:inline-block;
  padding:6px 10px;
  border-radius:999px;
  border:1px solid var(--line);
  background:#fff;
  color:#143b8f;
  font-size:12px;
  font-weight:700;
  margin-right:6px;
}
</style>
"""


def initialize_state():
    if "uploader_key" not in st.session_state:
        st.session_state["uploader_key"] = 0


def safe_float(value):
    if value in (None, ""):
        return None
    try:
        return float(value)
    except Exception:
        return None


def pct(value):
    value = safe_float(value)
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    return round(value * 100.0, 1)


def format_pct(value):
    p = pct(value)
    return "-" if p is None else f"{p:.1f}%"


def normalize_text(text):
    text = (text or "").strip().lower()
    repl = {
        "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u",
        "ü": "u", "ñ": "n"
    }
    for a, b in repl.items():
        text = text.replace(a, b)
    return " ".join(text.split())


def canonical_species_label(raw):
    n = normalize_text(raw)
    if "avic" in n:
        return "Avicultura"
    if "porc" in n:
        return "Porcino"
    if "conej" in n or "cun" in n:
        return "Conejos"
    if "caball" in n or "equin" in n:
        return "Caballos"
    if "vac" in n and ("lech" in n or "milk" in n):
        return "Vacuno leche"
    if "vac" in n and ("carne" in n or "beef" in n):
        return "Vacuno carne"
    if "ovin" in n or "caprin" in n or "peque" in n:
        return "Ovino y caprino"
    if "rumiante" in n and "vac" not in n:
        return "Ovino y caprino"
    return str(raw).strip() if raw not in (None, "") else "No disponible"


def species_family(species_label):
    if species_label in {"Avicultura", "Porcino", "Conejos", "Caballos"}:
        return "Monogástricos"
    if species_label in {"Vacuno leche", "Vacuno carne", "Ovino y caprino"}:
        return "Rumiantes"
    return "No disponible"


def canonical_trunk_name(name):
    n = normalize_text(name)
    if "nutric" in n or "aliment" in n:
        return "Alimentación"
    if "sanidad" in n or "patolog" in n:
        return "Sanidad"
    if "manejo" in n:
        return "Manejo"
    if "herramient" in n:
        return "Herramientas"
    return str(name).strip() if name not in (None, "") else "General"


def strongest_and_weakest_trunk(trunks_df):
    if trunks_df.empty:
        return "-", "-"
    tmp = trunks_df.dropna(subset=["vs_goal"]).copy()
    if tmp.empty:
        return "-", "-"
    tmp["canon"] = tmp["tronco"].apply(canonical_trunk_name)
    return (
        tmp.sort_values("vs_goal", ascending=False).iloc[0]["canon"],
        tmp.sort_values("vs_goal", ascending=True).iloc[0]["canon"],
    )


def get_asset_path(filename):
    for root in [Path("."), Path("/mnt/data")]:
        p = root / filename
        if p.exists():
            return p
    return None


ASSETS = {
    "nutreco": get_asset_path("Logo Nutreco.jpg"),
    "techteam": get_asset_path("Logo TechTeam 2.jpg"),
    "strip": get_asset_path("Solapa rosa.jpg"),
}


def image_data_uri(path):
    if not path or not path.exists():
        return ""
    data = path.read_bytes()
    mime = "image/jpeg" if path.suffix.lower() in [".jpg", ".jpeg"] else "image/png"
    return f"data:{mime};base64," + base64.b64encode(data).decode("ascii")


def display_corporate_header():
    if ASSETS["strip"]:
        st.image(str(ASSETS["strip"]), use_container_width=True)

    c1, c2, c3 = st.columns([1, 1.7, 1.25])
    with c1:
        if ASSETS["nutreco"]:
            st.image(str(ASSETS["nutreco"]), width=240)
    with c2:
        st.markdown(
            """
            <div class="corp-hero">
              <div style="font-size:34px;font-weight:800;line-height:1.05;">TechTeam · Gestión de desarrollo técnico de equipos</div>
              <div style="margin-top:10px;font-size:15px;line-height:1.5;">
                Fotografía comparativa del equipo técnico, detección de fortalezas y gaps,
                y propuesta de desarrollo equilibrado por áreas de assessment.
              </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with c3:
        if ASSETS["techteam"]:
            st.image(str(ASSETS["techteam"]), use_container_width=True)


def require_password():
    if st.session_state.get("auth_ok"):
        return
    st.markdown(STYLE, unsafe_allow_html=True)
    display_corporate_header()
    st.markdown("### Acceso restringido")
    st.write("Introduce la contraseña corporativa para acceder a la herramienta.")
    pwd = st.text_input("Contraseña", type="password")
    c1, c2 = st.columns([1, 4])
    with c1:
        if st.button("Acceder", type="primary", use_container_width=True):
            if pwd == APP_PASSWORD:
                st.session_state["auth_ok"] = True
                st.session_state.pop("auth_error", None)
                st.rerun()
            else:
                st.session_state["auth_error"] = "Contraseña incorrecta."
    with c2:
        st.caption("La contraseña se solicita de nuevo al iniciar cada sesión del navegador.")
    if st.session_state.get("auth_error"):
        st.error(st.session_state["auth_error"])
    st.stop()


def build_indicator_frame(ref_ws, eval_ws):
    rows = []
    current_area = None

    for ref_row, eval_row in zip(range(4, 29), range(9, 34)):
        area = ref_ws[f"B{ref_row}"].value
        if area:
            current_area = str(area).strip()

        indicator = ref_ws[f"C{ref_row}"].value
        if not indicator:
            continue
        indicator = str(indicator).strip()

        weight = safe_float(ref_ws[f"D{ref_row}"].value) or 0.0
        objective_raw = safe_float(ref_ws[f"E{ref_row}"].value) or 0.0
        objective_weighted = safe_float(ref_ws[f"F{ref_row}"].value)
        max_weighted = safe_float(ref_ws[f"G{ref_row}"].value)
        bbdd_raw = safe_float(ref_ws[f"H{ref_row}"].value)
        bbdd_weighted = safe_float(ref_ws[f"I{ref_row}"].value)

        raw_score = safe_float(eval_ws[f"D{eval_row}"].value)
        weighted_score = None if raw_score is None else raw_score * weight

        if objective_weighted is None:
            objective_weighted = weight * objective_raw
        if max_weighted is None:
            max_weighted = weight * 4
        if bbdd_weighted is None and bbdd_raw is not None:
            bbdd_weighted = weight * bbdd_raw

        rows.append(
            {
                "tronco": current_area,
                "indicator": indicator,
                "weight": weight,
                "score_raw": raw_score,
                "score_weighted": weighted_score,
                "objective_raw": objective_raw,
                "objective_weighted": objective_weighted,
                "max_weighted": max_weighted,
                "bbdd_raw": bbdd_raw,
                "bbdd_weighted": bbdd_weighted,
                "vs_goal": (weighted_score / objective_weighted) if objective_weighted and weighted_score is not None else None,
                "vs_max": (weighted_score / max_weighted) if max_weighted and weighted_score is not None else None,
                "vs_bbdd": (weighted_score / bbdd_weighted) if bbdd_weighted and weighted_score is not None else None,
            }
        )
    return pd.DataFrame(rows)


def summarise_trunks(indicators_df):
    rows = []
    for trunk, grp in indicators_df.groupby("tronco", dropna=False):
        score_weighted = grp["score_weighted"].sum(skipna=True)
        objective_weighted = grp["objective_weighted"].sum(skipna=True)
        max_weighted = grp["max_weighted"].sum(skipna=True)
        bbdd_weighted = grp["bbdd_weighted"].sum(skipna=True)
        rows.append(
            {
                "tronco": canonical_trunk_name(trunk),
                "score_raw_avg": grp["score_raw"].mean(skipna=True),
                "score_weighted_total": score_weighted,
                "objective_weighted_total": objective_weighted,
                "max_weighted_total": max_weighted,
                "bbdd_weighted_total": bbdd_weighted,
                "vs_goal": (score_weighted / objective_weighted) if objective_weighted else None,
                "vs_max": (score_weighted / max_weighted) if max_weighted else None,
                "vs_bbdd": (score_weighted / bbdd_weighted) if bbdd_weighted else None,
            }
        )
    out = pd.DataFrame(rows)
    if not out.empty:
        out["tronco"] = pd.Categorical(out["tronco"], categories=TRUNK_ORDER, ordered=True)
        out = out.sort_values("tronco").reset_index(drop=True)
        out["tronco"] = out["tronco"].astype(str)
    return out


def summarise_global(indicators_df):
    score_weighted = indicators_df["score_weighted"].sum(skipna=True)
    objective_weighted = indicators_df["objective_weighted"].sum(skipna=True)
    max_weighted = indicators_df["max_weighted"].sum(skipna=True)
    bbdd_weighted = indicators_df["bbdd_weighted"].sum(skipna=True)
    return {
        "score_raw_avg": indicators_df["score_raw"].mean(skipna=True),
        "vs_goal": (score_weighted / objective_weighted) if objective_weighted else None,
        "vs_max": (score_weighted / max_weighted) if max_weighted else None,
        "vs_bbdd": (score_weighted / bbdd_weighted) if bbdd_weighted else None,
    }


def extract_excel_level(wb_values, global_summary):
    ws = wb_values["ENLACES DATOS"] if "ENLACES DATOS" in wb_values.sheetnames else None
    if ws is None:
        return {"rank": None, "label": "-", "ratio": None, "cuts": []}

    ratio = safe_float(ws["K37"].value)
    if ratio is None:
        ratio = safe_float(global_summary["vs_max"])

    cuts = []
    for r in range(56, 62):
        cutoff = safe_float(ws[f"M{r}"].value)
        rank = safe_float(ws[f"O{r}"].value)
        label = ws[f"P{r}"].value
        q_flag = safe_float(ws[f"Q{r}"].value)
        if cutoff is None or rank is None or label in (None, ""):
            continue
        cuts.append({
            "cutoff": cutoff,
            "rank": int(rank),
            "label": str(label).strip(),
            "q_flag": int(q_flag) if q_flag is not None else None,
        })

    cuts = sorted(cuts, key=lambda x: x["cutoff"], reverse=True)
    rank_value = safe_float(ws["M44"].value)
    label_value = ws["N44"].value

    if rank_value is None and ratio is not None:
        rank_value = sum(1 for c in cuts if ratio > c["cutoff"])
    if label_value in (None, "") and rank_value is not None:
        for c in cuts:
            if c["rank"] == int(rank_value):
                label_value = c["label"]
                break

    return {
        "rank": int(rank_value) if rank_value is not None else None,
        "label": str(label_value).strip() if label_value not in (None, "") else "-",
        "ratio": ratio,
        "cuts": cuts,
    }


def parse_candidate(uploaded_file):
    suffix = Path(uploaded_file.name).suffix or ".xlsm"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(uploaded_file.getvalue())
        temp_path = tmp.name

    wb_values = load_workbook(temp_path, data_only=True, keep_vba=True)
    required = {"REFERENCIAS", "EVALUACION"}
    if not required.issubset(set(wb_values.sheetnames)):
        raise ValueError(f"{uploaded_file.name} no contiene las hojas mínimas esperadas: REFERENCIAS y EVALUACION.")

    ref_ws = wb_values["REFERENCIAS"]
    eval_ws = wb_values["EVALUACION"]
    mod_ws = wb_values["MODULO"] if "MODULO" in wb_values.sheetnames else None

    name = eval_ws["C2"].value or (mod_ws["H14"].value if mod_ws else None) or uploaded_file.name
    species = eval_ws["C4"].value or (mod_ws["H15"].value if mod_ws else None) or "-"
    date_value = eval_ws["C6"].value or (mod_ws["H16"].value if mod_ws else None)

    indicators_df = build_indicator_frame(ref_ws, eval_ws)
    if indicators_df["score_raw"].notna().sum() < 20:
        raise ValueError(f"{uploaded_file.name} no parece tener suficientes puntuaciones válidas. Revisa el Excel y vuelve a subirlo.")

    trunks_df = summarise_trunks(indicators_df)
    global_summary = summarise_global(indicators_df)
    strongest_trunk, weakest_trunk = strongest_and_weakest_trunk(trunks_df)
    excel_level = extract_excel_level(wb_values, global_summary)

    return {
        "filename": uploaded_file.name,
        "name": str(name).strip(),
        "species_raw": str(species).strip() if species is not None else "-",
        "species": canonical_species_label(species),
        "family": species_family(canonical_species_label(species)),
        "date": str(date_value) if date_value is not None else "No disponible",
        "indicators": indicators_df,
        "trunks": trunks_df,
        "global": global_summary,
        "global_level": excel_level["label"],
        "excel_rank": excel_level["rank"],
        "excel_ratio": excel_level["ratio"],
        "excel_cuts": excel_level["cuts"],
        "strongest_trunk": strongest_trunk,
        "weakest_trunk": weakest_trunk,
    }


def top_strengths(indicators_df, top_n=3):
    tmp = indicators_df.copy()
    tmp["composite"] = tmp["vs_goal"].fillna(0) * 0.65 + tmp["vs_bbdd"].fillna(0) * 0.20 + tmp["vs_max"].fillna(0) * 0.15
    top = tmp.sort_values("composite", ascending=False).head(top_n)
    return top[["indicator", "tronco", "score_raw", "vs_goal", "vs_bbdd"]].to_dict("records")


def main_gaps(indicators_df, top_n=3):
    tmp = indicators_df.copy()
    tmp["composite"] = tmp["vs_goal"].fillna(0) * 0.70 + tmp["vs_bbdd"].fillna(0) * 0.30
    low = tmp.sort_values("composite", ascending=True).head(top_n)
    return low[["indicator", "tronco", "score_raw", "vs_goal", "vs_bbdd"]].to_dict("records")


def filter_candidates(candidates, family_focus, species_focus):
    out = []
    excluded = []
    for c in candidates:
        if c["family"] != family_focus:
            excluded.append(c["name"])
            continue
        if species_focus.startswith("Equipo mixto"):
            out.append(c)
        elif c["species"] == species_focus:
            out.append(c)
        else:
            excluded.append(c["name"])
    return out, excluded


def level_distribution(candidates):
    rows = [{"Nivel": c["global_level"], "Nombre": c["name"]} for c in candidates]
    df = pd.DataFrame(rows)
    if df.empty:
        return pd.DataFrame(columns=["Nivel", "N"])
    order = ["Básico", "Basico", "Controla", "Supera", "Certificado", "Excelente", "Máster", "Master"]
    out = df.groupby("Nivel", as_index=False).size().rename(columns={"size": "N"})
    out["order"] = out["Nivel"].apply(lambda x: order.index(x) if x in order else 999)
    out = out.sort_values("order").drop(columns=["order"])
    return out


def build_team_trunk_summary(candidates):
    rows = []
    for trunk in TRUNK_ORDER:
        members = []
        for c in candidates:
            row = c["trunks"][c["trunks"]["tronco"] == trunk]
            if not row.empty:
                members.append({
                    "name": c["name"],
                    "species": c["species"],
                    "vs_goal": safe_float(row.iloc[0]["vs_goal"]),
                    "vs_bbdd": safe_float(row.iloc[0]["vs_bbdd"]),
                    "raw_avg": safe_float(row.iloc[0]["score_raw_avg"]),
                    "global_level": c["global_level"],
                })
        if not members:
            continue

        member_df = pd.DataFrame(members)
        mean_vs_goal = member_df["vs_goal"].dropna().mean()
        mean_vs_bbdd = member_df["vs_bbdd"].dropna().mean()
        strong_refs = member_df[member_df["vs_goal"].fillna(0) >= 1.0].sort_values("vs_goal", ascending=False)
        backup_pool = member_df[(member_df["vs_goal"].fillna(0) >= 0.85)].sort_values("vs_goal", ascending=False)

        status = "Gap crítico"
        if mean_vs_goal is not None and not math.isnan(mean_vs_goal):
            if mean_vs_goal >= 1.0 and len(strong_refs) >= 2:
                status = "Fortaleza consolidada"
            elif mean_vs_goal >= 0.95 and len(strong_refs) >= 1:
                status = "Aceptable con riesgo"
            elif mean_vs_goal >= 0.85:
                status = "Gap moderado"

        leaders = ", ".join(strong_refs["name"].head(2).tolist()) or ", ".join(member_df.sort_values("vs_goal", ascending=False)["name"].head(2).tolist())
        developers = ", ".join(
            backup_pool[~backup_pool["name"].isin(strong_refs["name"])].sort_values("vs_goal", ascending=False)["name"].head(2).tolist()
        )
        if not developers:
            developers = ", ".join(member_df[~member_df["name"].isin(strong_refs["name"])].sort_values("vs_goal", ascending=False)["name"].head(2).tolist())

        concentration_risk = "Bajo"
        if len(strong_refs) == 0:
            concentration_risk = "Alto"
        elif len(strong_refs) == 1:
            concentration_risk = "Medio"

        priority = (1 - (mean_vs_goal if mean_vs_goal is not None and not math.isnan(mean_vs_goal) else 0)) * 100
        if concentration_risk == "Alto":
            priority += 15
        elif concentration_risk == "Medio":
            priority += 8

        rows.append({
            "Tronco": trunk,
            "Media vs objetivo": round(mean_vs_goal, 3) if mean_vs_goal is not None and not math.isnan(mean_vs_goal) else None,
            "Media vs BBDD": round(mean_vs_bbdd, 3) if mean_vs_bbdd is not None and not math.isnan(mean_vs_bbdd) else None,
            "Referentes sólidos": int(len(strong_refs)),
            "Cobertura actual": leaders,
            "Personas a desarrollar": developers,
            "Estado equipo": status,
            "Riesgo de dependencia": concentration_risk,
            "Prioridad": round(priority, 1),
        })
    out = pd.DataFrame(rows)
    if not out.empty:
        out = out.sort_values(["Prioridad", "Tronco"], ascending=[False, True]).reset_index(drop=True)
    return out


def build_team_indicator_gaps(candidates, top_n=8):
    rows = []
    for c in candidates:
        tmp = c["indicators"].copy()
        tmp["name"] = c["name"]
        tmp["species"] = c["species"]
        tmp["tronco_canon"] = tmp["tronco"].apply(canonical_trunk_name)
        rows.append(tmp)
    if not rows:
        return pd.DataFrame()
    all_ind = pd.concat(rows, ignore_index=True)
    grouped = all_ind.groupby(["tronco_canon", "indicator"], as_index=False).agg(
        team_mean_vs_goal=("vs_goal", "mean"),
        team_mean_vs_bbdd=("vs_bbdd", "mean"),
        n_people=("name", "nunique"),
    )
    leaders = (
        all_ind.sort_values(["indicator", "vs_goal"], ascending=[True, False])
        .groupby(["tronco_canon", "indicator"], as_index=False)
        .agg(
            leader_1=("name", lambda s: s.iloc[0] if len(s) > 0 else ""),
            leader_2=("name", lambda s: s.iloc[1] if len(s) > 1 else ""),
        )
    )
    grouped = grouped.merge(leaders, on=["tronco_canon", "indicator"], how="left")
    grouped["gap_priority"] = (1 - grouped["team_mean_vs_goal"].fillna(0)) * 100 + (1 - grouped["team_mean_vs_bbdd"].fillna(0)) * 25
    grouped = grouped.sort_values("gap_priority", ascending=False).head(top_n).reset_index(drop=True)
    grouped["Referentes actuales"] = grouped.apply(
        lambda r: ", ".join([x for x in [r["leader_1"], r["leader_2"]] if x]), axis=1
    )
    grouped = grouped.rename(columns={
        "tronco_canon": "Tronco",
        "indicator": "Indicador",
        "team_mean_vs_goal": "Media equipo vs objetivo",
        "team_mean_vs_bbdd": "Media equipo vs BBDD",
        "n_people": "N técnicos",
    })
    return grouped[["Tronco", "Indicador", "Media equipo vs objetivo", "Media equipo vs BBDD", "N técnicos", "Referentes actuales", "gap_priority"]]


def build_candidate_trunk_matrix(candidates):
    rows = []
    for c in candidates:
        row = {"Nombre": c["name"], "Especie": c["species"], "Calificación Excel": c["global_level"]}
        for trunk in TRUNK_ORDER:
            df = c["trunks"][c["trunks"]["tronco"] == trunk]
            row[trunk] = safe_float(df.iloc[0]["vs_goal"]) if not df.empty else None
        rows.append(row)
    return pd.DataFrame(rows)


def build_team_development_plan(team_trunk_summary, candidates):
    lines = []
    top_gaps = team_trunk_summary.sort_values("Prioridad", ascending=False).head(3)
    top_strengths = team_trunk_summary.sort_values("Prioridad", ascending=True).head(2)

    lines.append("## Resumen directivo")
    lines.append(
        "El objetivo del plan no es jerarquizar a los técnicos, sino equilibrar el equipo, reducir dependencias y asegurar cobertura suficiente en los cuatro troncos del assessment."
    )
    lines.append("")

    lines.append("## Prioridades de desarrollo del equipo")
    if top_gaps.empty:
        lines.append("- No se han detectado gaps críticos con la selección actual.")
    else:
        for _, row in top_gaps.iterrows():
            lines.append(
                f"- {row['Tronco']}: estado '{row['Estado equipo']}', media {format_pct(row['Media vs objetivo'])}, "
                f"cobertura actual: {row['Cobertura actual']}. Personas a desarrollar: {row['Personas a desarrollar']}."
            )
    lines.append("")

    lines.append("## Fortalezas a aprovechar")
    if top_strengths.empty:
        lines.append("- No hay suficientes datos para identificar fortalezas de equipo.")
    else:
        for _, row in top_strengths.iterrows():
            lines.append(
                f"- {row['Tronco']}: puede actuar como palanca interna. Referentes actuales: {row['Cobertura actual']}."
            )
    lines.append("")

    lines.append("## Plan de acción recomendado")
    priority_blocks = []
    for _, row in top_gaps.iterrows():
        priority_blocks.append(
            {
                "tronco": row["Tronco"],
                "goal": row["Media vs objetivo"],
                "leaders": row["Cobertura actual"],
                "developers": row["Personas a desarrollar"],
                "status": row["Estado equipo"],
                "risk": row["Riesgo de dependencia"],
            }
        )

    if not priority_blocks:
        lines.append("- Mantener el equilibrio actual y usar a los referentes por área como mentores internos.")
    else:
        for idx, block in enumerate(priority_blocks, start=1):
            lines.append(f"### Prioridad {idx} · {block['tronco']}")
            lines.append(
                f"- Objetivo: elevar la cobertura del equipo en {block['tronco']} hasta, al menos, 100% frente al objetivo y disponer de dos personas sólidas en esta área."
            )
            lines.append(f"- Referentes actuales: {block['leaders'] or 'No disponible'}.")
            lines.append(f"- Personas a desarrollar: {block['developers'] or 'No disponible'}.")
            lines.append(
                f"- Acción 0-3 meses: sesiones de revisión de casos reales, visitas conjuntas y transferencia estructurada desde los referentes actuales."
            )
            lines.append(
                f"- Acción 3-6 meses: asignación de formación específica, revisión de indicadores del tronco y seguimiento del progreso en evaluaciones intermedias."
            )
            lines.append(
                f"- Riesgo actual: {block['risk']}. Estado del equipo en este tronco: {block['status']}."
            )
    return "\n".join(lines)


def team_summary_metrics(candidates, team_trunk_summary):
    globals_df = pd.DataFrame(
        [
            {
                "name": c["name"],
                "vs_goal": c["global"]["vs_goal"],
                "vs_bbdd": c["global"]["vs_bbdd"],
                "level": c["global_level"],
            }
            for c in candidates
        ]
    )
    avg_vs_goal = globals_df["vs_goal"].dropna().mean() if not globals_df.empty else None
    avg_vs_bbdd = globals_df["vs_bbdd"].dropna().mean() if not globals_df.empty else None
    balance_index = None
    if not team_trunk_summary.empty:
        vals = team_trunk_summary["Media vs objetivo"].dropna().tolist()
        if vals:
            balance_index = max(0.0, min(100.0, 100 - (pd.Series(vals).std(ddof=0) * 100)))
    n_critical = 0 if team_trunk_summary.empty else int((team_trunk_summary["Estado equipo"] == "Gap crítico").sum())
    return {
        "n_people": len(candidates),
        "avg_vs_goal": avg_vs_goal,
        "avg_vs_bbdd": avg_vs_bbdd,
        "balance_index": balance_index,
        "n_critical": n_critical,
    }


def build_levels_chart(level_df):
    if level_df.empty:
        return None
    fig = px.bar(level_df, x="Nivel", y="N", text="N")
    fig.update_traces(textposition="outside")
    fig.update_layout(height=380, xaxis_title="", yaxis_title="Nº técnicos", showlegend=False, margin=dict(l=20, r=20, t=20, b=20))
    return fig


def build_heatmap(candidate_trunk_df):
    if candidate_trunk_df.empty:
        return None

    # En Streamlit Cloud puede haber versiones de pandas donde applymap genere problemas
    # de compatibilidad. Además, si falta algún tronco en un caso concreto, reindex evita
    # errores por columnas ausentes.
    base_df = candidate_trunk_df.set_index("Nombre").reindex(columns=TRUNK_ORDER)
    plot_df = base_df.apply(lambda col: col.map(lambda x: pct(x) if x is not None and not pd.isna(x) else None))

    # Si todo está vacío, no intentamos dibujar el heatmap.
    if plot_df.dropna(how="all").empty:
        return None

    fig = px.imshow(
        plot_df,
        aspect="auto",
        text_auto=".0f",
        color_continuous_scale="RdYlGn",
        origin="lower",
        labels={"x": "Tronco", "y": "Técnico", "color": "% vs objetivo"},
    )
    fig.update_layout(height=max(420, 70 * len(plot_df)), margin=dict(l=20, r=20, t=20, b=20))
    return fig


def build_team_trunk_chart(team_trunk_summary):
    if team_trunk_summary.empty:
        return None
    plot_df = team_trunk_summary.copy()
    plot_df["Media %"] = plot_df["Media vs objetivo"].apply(lambda x: pct(x) if x is not None else None)
    fig = px.bar(plot_df, x="Tronco", y="Media %", text="Media %", color="Estado equipo")
    fig.update_traces(textposition="outside")
    fig.update_layout(height=400, xaxis_title="", yaxis_title="% vs objetivo", margin=dict(l=20, r=20, t=20, b=20))
    return fig


def build_radar_comparison(candidates, selected_names):
    if not candidates:
        return None
    fig = go.Figure()
    used = [c for c in candidates if c["name"] in selected_names] if selected_names else candidates[:6]
    for c in used:
        values = []
        for trunk in TRUNK_ORDER:
            df = c["trunks"][c["trunks"]["tronco"] == trunk]
            values.append((safe_float(df.iloc[0]["vs_goal"]) or 0) * 100 if not df.empty else 0)
        fig.add_trace(
            go.Scatterpolar(
                r=values + [values[0]],
                theta=TRUNK_ORDER + [TRUNK_ORDER[0]],
                fill="toself",
                name=c["name"],
                opacity=0.18,
                line=dict(width=2),
            )
        )
    fig.update_layout(
        polar=dict(radialaxis=dict(visible=True, range=[0, 140])),
        showlegend=True,
        height=620,
        margin=dict(l=20, r=20, t=20, b=20),
    )
    return fig


def figure_to_png_bytes(fig):
    if fig is None:
        return None
    try:
        return fig.to_image(format="png", width=1200, height=800, scale=2)
    except Exception:
        return None


def markdownish_to_html(report_text):
    parts = []
    in_list = False
    for raw in report_text.splitlines():
        line = raw.strip()
        if not line:
            if in_list:
                parts.append("</ul>")
                in_list = False
            continue
        if line.startswith("## "):
            if in_list:
                parts.append("</ul>")
                in_list = False
            parts.append(f"<h2>{escape(line[3:])}</h2>")
        elif line.startswith("### "):
            if in_list:
                parts.append("</ul>")
                in_list = False
            parts.append(f"<h3>{escape(line[4:])}</h3>")
        elif line.startswith("- "):
            if not in_list:
                parts.append("<ul>")
                in_list = True
            parts.append(f"<li>{escape(line[2:])}</li>")
        else:
            if in_list:
                parts.append("</ul>")
                in_list = False
            parts.append(f"<p>{escape(line)}</p>")
    if in_list:
        parts.append("</ul>")
    return "\n".join(parts)


def chart_images_html(chart_images):
    blocks = []
    for title, image_bytes in chart_images:
        if not image_bytes:
            continue
        b64 = base64.b64encode(image_bytes).decode("ascii")
        blocks.append(
            f"<div class='chart-card'><h3>{escape(title)}</h3><img src='data:image/png;base64,{b64}' alt='{escape(title)}'></div>"
        )
    return "".join(blocks)


def build_html_report(candidates, family_focus, species_focus, team_metrics, team_trunk_summary, indicator_gaps, report_text, chart_images):
    nutreco_uri = image_data_uri(ASSETS["nutreco"])
    techteam_uri = image_data_uri(ASSETS["techteam"])
    strip_uri = image_data_uri(ASSETS["strip"])

    candidate_rows = []
    for c in sorted(candidates, key=lambda x: x["name"]):
        candidate_rows.append(
            f"<tr><td>{escape(c['name'])}</td><td>{escape(c['species'])}</td><td>{escape(c['global_level'])}</td><td>{format_pct(c['global']['vs_goal'])}</td><td>{escape(c['strongest_trunk'])}</td><td>{escape(c['weakest_trunk'])}</td></tr>"
        )

    trunk_rows = []
    for _, r in team_trunk_summary.iterrows():
        trunk_rows.append(
            f"<tr><td>{escape(str(r['Tronco']))}</td><td>{format_pct(r['Media vs objetivo'])}</td><td>{format_pct(r['Media vs BBDD'])}</td><td>{escape(str(r['Estado equipo']))}</td><td>{escape(str(r['Cobertura actual']))}</td><td>{escape(str(r['Personas a desarrollar']))}</td></tr>"
        )

    gap_rows = []
    for _, r in indicator_gaps.iterrows():
        gap_rows.append(
            f"<tr><td>{escape(str(r['Tronco']))}</td><td>{escape(str(r['Indicador']))}</td><td>{format_pct(r['Media equipo vs objetivo'])}</td><td>{escape(str(r['Referentes actuales']))}</td></tr>"
        )

    body_html = markdownish_to_html(report_text)
    charts_html = chart_images_html(chart_images)

    html_doc = f"""<!DOCTYPE html>
<html lang='es'>
<head>
<meta charset='utf-8'>
<meta name='viewport' content='width=device-width, initial-scale=1'>
<title>Informe corporativo · Desarrollo técnico de equipos</title>
<style>
:root {{
  --nutreco-blue:#143b8f;
  --pink:#d81b90;
  --red:#ef233c;
  --line:#dbe3ef;
  --muted:#6b7280;
}}
* {{ box-sizing:border-box; }}
body {{ margin:0; font-family:Arial, Helvetica, sans-serif; color:#1f2937; background:#f6f8fc; }}
.top-strip {{ height:28px; background-image:url('{strip_uri}'); background-size:cover; background-position:center; }}
.container {{ max-width:1240px; margin:0 auto; padding:22px; }}
.hero {{
  background:linear-gradient(135deg, rgba(20,59,143,.97), rgba(216,27,144,.94));
  color:#fff; border-radius:22px; padding:18px 20px;
}}
.hero-grid {{ display:grid; grid-template-columns:240px 1fr 300px; gap:18px; align-items:center; }}
.hero-grid img {{ max-width:100%; max-height:120px; object-fit:contain; background:#fff; border-radius:16px; padding:10px; }}
.hero h1 {{ margin:0 0 8px; font-size:30px; line-height:1.05; }}
.hero p {{ margin:0; font-size:15px; line-height:1.5; }}
.metrics {{ display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:14px; margin-top:18px; }}
.metric {{ background:#fff; border:1px solid var(--line); border-radius:16px; padding:14px; }}
.metric small {{ display:block; color:var(--muted); margin-bottom:6px; }}
.metric strong {{ color:var(--nutreco-blue); font-size:20px; }}
.section {{ background:#fff; border:1px solid var(--line); border-radius:18px; padding:18px; margin-top:18px; }}
.section h2 {{ color:var(--nutreco-blue); font-size:20px; margin:0 0 10px; }}
.section h3 {{ color:#111827; font-size:16px; margin:18px 0 8px; }}
.section p, .section li {{ font-size:14px; line-height:1.55; }}
table {{ width:100%; border-collapse:collapse; margin-top:10px; }}
th, td {{ border:1px solid var(--line); padding:8px; font-size:13px; text-align:left; vertical-align:top; }}
th {{ background:#f8fafc; }}
.chart-grid {{ display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:16px; margin-top:16px; }}
.chart-card {{ border:1px solid var(--line); border-radius:16px; padding:12px; }}
.chart-card img {{ width:100%; height:auto; display:block; }}
.footer {{ margin-top:18px; color:var(--muted); font-size:12px; }}
@media (max-width:1000px) {{
  .hero-grid, .metrics, .chart-grid {{ grid-template-columns:1fr; }}
}}
</style>
</head>
<body>
<div class='top-strip'></div>
<div class='container'>
  <div class='hero'>
    <div class='hero-grid'>
      <img src='{nutreco_uri}' alt='Nutreco'>
      <div>
        <h1>Informe corporativo · Gestión de desarrollo técnico de equipos</h1>
        <p>Fotografía comparativa del equipo, cobertura por troncos del assessment y plan de desarrollo para construir un equipo equilibrado.</p>
      </div>
      <img src='{techteam_uri}' alt='TechTeam'>
    </div>
  </div>

  <div class='metrics'>
    <div class='metric'><small>Bloque</small><strong>{escape(family_focus)}</strong></div>
    <div class='metric'><small>Especie / subespecie</small><strong>{escape(species_focus)}</strong></div>
    <div class='metric'><small>Nº técnicos analizados</small><strong>{team_metrics['n_people']}</strong></div>
    <div class='metric'><small>Gaps críticos de equipo</small><strong>{team_metrics['n_critical']}</strong></div>
  </div>
  <div class='metrics'>
    <div class='metric'><small>Media equipo vs objetivo</small><strong>{format_pct(team_metrics['avg_vs_goal'])}</strong></div>
    <div class='metric'><small>Media equipo vs BBDD</small><strong>{format_pct(team_metrics['avg_vs_bbdd'])}</strong></div>
    <div class='metric'><small>Índice de equilibrio</small><strong>{'-' if team_metrics['balance_index'] is None else f"{team_metrics['balance_index']:.1f}/100"}</strong></div>
    <div class='metric'><small>Fecha</small><strong>{datetime.now().strftime('%d/%m/%Y')}</strong></div>
  </div>

  <div class='section'>
    <h2>Resumen ejecutivo</h2>
    {body_html}
  </div>

  <div class='section'>
    <h2>Equipo analizado</h2>
    <table>
      <thead><tr><th>Técnico</th><th>Especie</th><th>Calificación Excel</th><th>Vs objetivo global</th><th>Tronco fuerte</th><th>Tronco débil</th></tr></thead>
      <tbody>{''.join(candidate_rows)}</tbody>
    </table>
  </div>

  <div class='section'>
    <h2>Cobertura de equipo por tronco</h2>
    <table>
      <thead><tr><th>Tronco</th><th>Media vs objetivo</th><th>Media vs BBDD</th><th>Estado</th><th>Cobertura actual</th><th>Personas a desarrollar</th></tr></thead>
      <tbody>{''.join(trunk_rows)}</tbody>
    </table>
  </div>

  <div class='section'>
    <h2>Indicadores prioritarios a reforzar</h2>
    <table>
      <thead><tr><th>Tronco</th><th>Indicador</th><th>Media equipo vs objetivo</th><th>Referentes actuales</th></tr></thead>
      <tbody>{''.join(gap_rows)}</tbody>
    </table>
  </div>

  <div class='section'>
    <h2>Gráficos comparativos</h2>
    <div class='chart-grid'>{charts_html}</div>
  </div>

  <div class='footer'>Documento interno corporativo · Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}</div>
</div>
</body>
</html>"""
    return html_doc.encode("utf-8")


def safe_rl_image(path_or_bytes, width=None, height=None):
    try:
        return RLImage(path_or_bytes, width=width, height=height)
    except Exception:
        return None


def build_pdf_report(candidates, family_focus, species_focus, team_metrics, team_trunk_summary, indicator_gaps, report_text, chart_images):
    bio = io.BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=A4, leftMargin=32, rightMargin=32, topMargin=26, bottomMargin=26)
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], fontName="Helvetica-Bold", fontSize=17, leading=20, textColor=colors.HexColor("#143b8f"))
    h2 = ParagraphStyle("H2", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=13, leading=16, textColor=colors.HexColor("#143b8f"))
    body = ParagraphStyle("Body", parent=styles["BodyText"], fontName="Helvetica", fontSize=9.4, leading=13)
    bullet = ParagraphStyle("Bullet", parent=body, leftIndent=14, bulletIndent=0)

    story = []

    strip = ASSETS["strip"]
    if strip:
        img = safe_rl_image(str(strip), width=530, height=22)
        if img:
            story.extend([img, Spacer(1, 8)])

    logos = []
    for key, width in [("nutreco", 120), ("techteam", 150)]:
        p = ASSETS[key]
        if p:
            img = safe_rl_image(str(p), width=width, height=55)
            if img:
                logos.append(img)
    if logos:
        data = [logos]
        tbl = Table(data)
        tbl.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
        story.extend([tbl, Spacer(1, 10)])

    story.append(Paragraph("Gestión de desarrollo técnico de equipos", h1))
    story.append(Spacer(1, 6))
    story.append(Paragraph(f"Bloque: {escape(family_focus)} · Especie/subespecie: {escape(species_focus)}", body))
    story.append(Spacer(1, 8))

    metrics_data = [
        ["N técnicos", str(team_metrics["n_people"]), "Media vs objetivo", format_pct(team_metrics["avg_vs_goal"])],
        ["Media vs BBDD", format_pct(team_metrics["avg_vs_bbdd"]), "Índice de equilibrio", "-" if team_metrics["balance_index"] is None else f"{team_metrics['balance_index']:.1f}/100"],
        ["Gaps críticos", str(team_metrics["n_critical"]), "Fecha", datetime.now().strftime("%d/%m/%Y")],
    ]
    mt = Table(metrics_data, colWidths=[90, 90, 110, 120])
    mt.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#dbe3ef")),
        ("BACKGROUND", (0,0), (-1,-1), colors.white),
        ("TEXTCOLOR", (0,0), (-1,-1), colors.HexColor("#1f2937")),
        ("FONTNAME", (0,0), (-1,-1), "Helvetica"),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("LEADING", (0,0), (-1,-1), 11),
    ]))
    story.extend([mt, Spacer(1, 12)])

    story.append(Paragraph("Resumen ejecutivo", h2))
    for raw in report_text.splitlines():
        line = raw.strip()
        if not line:
            story.append(Spacer(1, 4))
        elif line.startswith("## "):
            story.append(Paragraph(escape(line[3:]), h2))
        elif line.startswith("### "):
            story.append(Paragraph(escape(line[4:]), body))
        elif line.startswith("- "):
            story.append(Paragraph("• " + escape(line[2:]), bullet))
        else:
            story.append(Paragraph(escape(line), body))
    story.append(Spacer(1, 10))

    story.append(Paragraph("Cobertura de equipo por tronco", h2))
    trunk_data = [["Tronco", "Media vs obj.", "Estado", "Cobertura actual", "Personas a desarrollar"]]
    for _, r in team_trunk_summary.iterrows():
        trunk_data.append([
            str(r["Tronco"]),
            format_pct(r["Media vs objetivo"]),
            str(r["Estado equipo"]),
            str(r["Cobertura actual"]),
            str(r["Personas a desarrollar"]),
        ])
    t1 = Table(trunk_data, colWidths=[80, 70, 90, 145, 145], repeatRows=1)
    t1.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#dbe3ef")),
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#f8fafc")),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 8.5),
        ("LEADING", (0,0), (-1,-1), 10),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
    ]))
    story.extend([t1, Spacer(1, 10)])

    story.append(Paragraph("Indicadores prioritarios a reforzar", h2))
    gap_data = [["Tronco", "Indicador", "Media vs obj.", "Referentes actuales"]]
    for _, r in indicator_gaps.head(6).iterrows():
        gap_data.append([
            str(r["Tronco"]),
            str(r["Indicador"]),
            format_pct(r["Media equipo vs objetivo"]),
            str(r["Referentes actuales"]),
        ])
    t2 = Table(gap_data, colWidths=[80, 180, 75, 165], repeatRows=1)
    t2.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#dbe3ef")),
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#f8fafc")),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 8.5),
        ("LEADING", (0,0), (-1,-1), 10),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
    ]))
    story.extend([t2, Spacer(1, 10)])

    if chart_images:
        story.append(Paragraph("Gráficos comparativos", h2))
        for title, img_bytes in chart_images:
            if img_bytes:
                story.append(Paragraph(escape(title), body))
                img = safe_rl_image(io.BytesIO(img_bytes), width=500, height=330)
                if img:
                    story.extend([img, Spacer(1, 8)])

    doc.build(story)
    bio.seek(0)
    return bio.getvalue()


def reset_uploads():
    st.session_state["uploader_key"] += 1
    st.rerun()


initialize_state()
require_password()
st.markdown(STYLE, unsafe_allow_html=True)
display_corporate_header()

st.title("TechTeam · Gestión de desarrollo técnico de equipos")
st.caption(
    "Compara técnicos sin jerarquizarlos globalmente, identifica fortalezas y gaps del equipo y propone un plan de desarrollo para construir un equipo equilibrado."
)

with st.expander("Cómo interpreta la herramienta el desarrollo del equipo", expanded=True):
    st.markdown(
        """
- Esta app **no busca elegir un senior** ni ordenar a unas personas por encima de otras.
- Su objetivo es construir una **fotografía técnica del equipo**.
- La herramienta compara resultados de assessment, fortalezas, debilidades y cobertura por áreas.
- A partir de esa comparación, propone:
  - qué troncos están fuertes como equipo;
  - qué gaps deben reforzarse;
  - quién puede actuar como referente interno por área;
  - y qué personas conviene desarrollar para reducir dependencia y equilibrar el equipo.
        """
    )

control1, control2, upload_col, clear_col = st.columns([1.2, 1.3, 3.2, 1.1])

with control1:
    family_focus = st.selectbox("Bloque productivo", ["Monogástricos", "Rumiantes"], index=0)
with control2:
    species_options = MONOGASTRIC_SPECIES if family_focus == "Monogástricos" else RUMINANT_SPECIES
    species_focus = st.selectbox("Especie / subespecie", species_options, index=0)
with upload_col:
    uploaded_files = st.file_uploader(
        "Sube de 2 a 15 evaluaciones (.xlsm o .xlsx)",
        type=["xlsm", "xlsx"],
        accept_multiple_files=True,
        key=f"uploader_{st.session_state['uploader_key']}",
        help="Pueden ser de una misma especie o un equipo mixto, pero siempre dentro del mismo bloque productivo.",
    )
with clear_col:
    st.write("")
    st.write("")
    if st.button("Borrar archivos", use_container_width=True):
        reset_uploads()

if not uploaded_files:
    st.stop()

if len(uploaded_files) < 2:
    st.warning("Necesitas subir al menos 2 evaluaciones para construir una fotografía de equipo.")
    st.stop()

if len(uploaded_files) > 15:
    st.warning("La app está pensada para un máximo de 15 evaluaciones por análisis.")
    st.stop()

parsed_candidates = []
errors = []
for file in uploaded_files:
    try:
        parsed_candidates.append(parse_candidate(file))
    except Exception as exc:
        errors.append(f"{file.name}: {exc}")

if errors:
    st.error("Algunos archivos no se han podido procesar:")
    for err in errors:
        st.write(f"- {err}")

if not parsed_candidates:
    st.stop()

selected_candidates, excluded_candidates = filter_candidates(parsed_candidates, family_focus, species_focus)

if excluded_candidates:
    st.warning(
        "Se han excluido del análisis algunos archivos por no encajar con el bloque o la especie/subespecie seleccionados: "
        + ", ".join(excluded_candidates)
    )

if len(selected_candidates) < 2:
    st.error("Con la selección actual no quedan al menos 2 técnicos comparables. Ajusta el selector de especie/subespecie o sube más archivos.")
    st.stop()

selected_candidates = sorted(selected_candidates, key=lambda x: x["name"].lower())

team_trunk_summary = build_team_trunk_summary(selected_candidates)
indicator_gaps = build_team_indicator_gaps(selected_candidates, top_n=8)
candidate_trunk_df = build_candidate_trunk_matrix(selected_candidates)
levels_df = level_distribution(selected_candidates)
team_metrics = team_summary_metrics(selected_candidates, team_trunk_summary)
report_text = build_team_development_plan(team_trunk_summary, selected_candidates)

tab1, tab2, tab3, tab4 = st.tabs([
    "Visión global del equipo",
    "Comparación técnica",
    "Cobertura y gaps",
    "Informe y descargas",
])

with tab1:
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Nº técnicos analizados", team_metrics["n_people"])
    m2.metric("Media equipo vs objetivo", format_pct(team_metrics["avg_vs_goal"]))
    m3.metric("Media equipo vs BBDD", format_pct(team_metrics["avg_vs_bbdd"]))
    m4.metric("Índice de equilibrio", "-" if team_metrics["balance_index"] is None else f"{team_metrics['balance_index']:.1f}/100")

    m5, m6, m7 = st.columns(3)
    m5.metric("Gaps críticos", int(team_metrics["n_critical"]))
    strongest_team = team_trunk_summary.sort_values("Prioridad", ascending=True).iloc[0]["Tronco"] if not team_trunk_summary.empty else "-"
    weakest_team = team_trunk_summary.sort_values("Prioridad", ascending=False).iloc[0]["Tronco"] if not team_trunk_summary.empty else "-"
    m6.metric("Palanca de equipo", strongest_team)
    m7.metric("Prioridad principal", weakest_team)

    st.markdown("### Lectura directiva")
    st.text_area(
        "Plan de desarrollo del equipo",
        value=report_text,
        height=360,
    )

    c1, c2 = st.columns([1.1, 1.2])
    with c1:
        st.markdown("### Distribución de niveles del equipo")
        fig_levels = build_levels_chart(levels_df)
        if fig_levels:
            st.plotly_chart(fig_levels, use_container_width=True)
    with c2:
        st.markdown("### Cobertura media por tronco")
        fig_team_trunks = build_team_trunk_chart(team_trunk_summary)
        if fig_team_trunks:
            st.plotly_chart(fig_team_trunks, use_container_width=True)

with tab2:
    st.markdown("### Matriz comparativa por troncos")
    candidate_view = candidate_trunk_df.copy()
    for trunk in TRUNK_ORDER:
        candidate_view[trunk] = candidate_view[trunk].apply(format_pct)
    st.dataframe(candidate_view, use_container_width=True, hide_index=True)

    st.markdown("### Heatmap del equipo vs objetivo")
    heatmap_fig = build_heatmap(build_candidate_trunk_matrix(selected_candidates))
    if heatmap_fig:
        st.plotly_chart(heatmap_fig, use_container_width=True)

    st.markdown("### Radar comparativo entre técnicos")
    radar_names = st.multiselect(
        "Selecciona técnicos para el radar comparativo",
        options=[c["name"] for c in selected_candidates],
        default=[c["name"] for c in selected_candidates[:min(6, len(selected_candidates))]],
    )
    radar_fig = build_radar_comparison(selected_candidates, radar_names)
    if radar_fig:
        st.plotly_chart(radar_fig, use_container_width=True)

    st.markdown("### Detalle por técnico")
    for c in selected_candidates:
        with st.expander(f"{c['name']} · {c['species']} · {c['global_level']}", expanded=False):
            left, right = st.columns([1, 1])
            with left:
                st.metric("Vs objetivo global", format_pct(c["global"]["vs_goal"]))
                st.metric("Vs BBDD global", format_pct(c["global"]["vs_bbdd"]))
                st.metric("Tronco más fuerte", c["strongest_trunk"])
                st.metric("Tronco más débil", c["weakest_trunk"])
            with right:
                strengths = top_strengths(c["indicators"], top_n=3)
                gaps = main_gaps(c["indicators"], top_n=3)
                st.markdown("**Fortalezas destacadas**")
                for item in strengths:
                    st.write(f"- {item['indicator']} ({canonical_trunk_name(item['tronco'])}) · {format_pct(item['vs_goal'])} vs objetivo")
                st.markdown("**Debilidades destacadas**")
                for item in gaps:
                    st.write(f"- {item['indicator']} ({canonical_trunk_name(item['tronco'])}) · {format_pct(item['vs_goal'])} vs objetivo")

            trunk_display = c["trunks"].copy()
            trunk_display["vs_goal"] = trunk_display["vs_goal"].apply(format_pct)
            trunk_display["vs_bbdd"] = trunk_display["vs_bbdd"].apply(format_pct)
            trunk_display["vs_max"] = trunk_display["vs_max"].apply(format_pct)
            trunk_display = trunk_display.rename(columns={
                "tronco": "Tronco",
                "score_raw_avg": "Score medio",
                "vs_goal": "Vs objetivo",
                "vs_bbdd": "Vs BBDD",
                "vs_max": "Vs máximo",
            })
            st.dataframe(trunk_display[["Tronco", "Score medio", "Vs objetivo", "Vs BBDD", "Vs máximo"]], use_container_width=True, hide_index=True)

with tab3:
    st.markdown("### Cobertura del equipo por tronco")
    coverage_display = team_trunk_summary.copy()
    coverage_display["Media vs objetivo"] = coverage_display["Media vs objetivo"].apply(format_pct)
    coverage_display["Media vs BBDD"] = coverage_display["Media vs BBDD"].apply(format_pct)
    st.dataframe(coverage_display, use_container_width=True, hide_index=True)

    st.markdown("### Indicadores más débiles del equipo")
    gap_display = indicator_gaps.copy()
    if not gap_display.empty:
        gap_display["Media equipo vs objetivo"] = gap_display["Media equipo vs objetivo"].apply(format_pct)
        gap_display["Media equipo vs BBDD"] = gap_display["Media equipo vs BBDD"].apply(format_pct)
        gap_display = gap_display.rename(columns={"gap_priority": "Prioridad"})
        st.dataframe(gap_display, use_container_width=True, hide_index=True)

    st.markdown("### Recomendación de cobertura y desarrollo")
    if team_trunk_summary.empty:
        st.info("No hay datos suficientes para construir la recomendación de cobertura.")
    else:
        for _, row in team_trunk_summary.iterrows():
            with st.expander(f"{row['Tronco']} · {row['Estado equipo']} · prioridad {row['Prioridad']:.1f}", expanded=row["Estado equipo"] in {"Gap crítico", "Gap moderado"}):
                st.write(f"**Cobertura actual:** {row['Cobertura actual'] or 'No disponible'}")
                st.write(f"**Personas a desarrollar:** {row['Personas a desarrollar'] or 'No disponible'}")
                st.write(f"**Riesgo de dependencia:** {row['Riesgo de dependencia']}")
                st.write(
                    "Se recomienda reforzar esta área con transferencia interna, acompañamiento en casos reales, revisión de indicadores específicos y formación dirigida."
                )

with tab4:
    st.markdown("### Informe corporativo")
    st.write("Puedes descargar un informe HTML o PDF con plantilla corporativa y logos.")

    fig_levels = build_levels_chart(levels_df)
    fig_team_trunks = build_team_trunk_chart(team_trunk_summary)
    heatmap_fig = build_heatmap(build_candidate_trunk_matrix(selected_candidates))
    radar_fig = build_radar_comparison(selected_candidates, [c["name"] for c in selected_candidates[:min(6, len(selected_candidates))]])
    chart_images = [
        ("Distribución de niveles del equipo", figure_to_png_bytes(fig_levels)),
        ("Cobertura media por tronco", figure_to_png_bytes(fig_team_trunks)),
        ("Heatmap del equipo vs objetivo", figure_to_png_bytes(heatmap_fig)),
        ("Radar comparativo entre técnicos", figure_to_png_bytes(radar_fig)),
    ]

    html_bytes = build_html_report(
        selected_candidates,
        family_focus,
        species_focus,
        team_metrics,
        team_trunk_summary,
        indicator_gaps,
        report_text,
        chart_images,
    )
    pdf_bytes = build_pdf_report(
        selected_candidates,
        family_focus,
        species_focus,
        team_metrics,
        team_trunk_summary,
        indicator_gaps,
        report_text,
        chart_images,
    )

    d1, d2 = st.columns(2)
    with d1:
        st.download_button(
            "Descargar informe HTML corporativo",
            data=html_bytes,
            file_name="informe_desarrollo_equipo_tecnico.html",
            mime="text/html",
            use_container_width=True,
        )
    with d2:
        st.download_button(
            "Descargar informe PDF corporativo",
            data=pdf_bytes,
            file_name="informe_desarrollo_equipo_tecnico.pdf",
            mime="application/pdf",
            use_container_width=True,
        )
